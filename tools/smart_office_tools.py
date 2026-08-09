"""
Smart Office Tools — end the all-nighter of manual data entry.

For someone who works in Microsoft Office in Arabic and English and retypes data
into Excel by hand: paste the data (or point at a file) and get a clean, formatted
spreadsheet in one step; and convert precisely between Excel · Word · PDF ·
PowerPoint · CSV.

  • data_to_excel   — raw pasted text OR a source file → a formatted .xlsx
  • office_convert  — precise conversion between Office formats (both languages)
  • extract_tables  — pull every table out of a PDF/Word into one Excel workbook
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Annotated

from langchain_core.tools import tool

from config import resolve_output_path
from core import office_convert as _oc


@tool
def data_to_excel(
    data: Annotated[str, "البيانات الخام (منسوخة كما هي) أو مسار ملف مصدر "
                         "(.csv/.txt/.tsv/.pdf/.docx). أعمدة مفصولة بفاصلة/تاب/فاصلة "
                         "منقوطة/مسافات — يُكتشف تلقائياً، عربي وإنجليزي."],
    dest: Annotated[str, "مسار ملف الإكسل الناتج (.xlsx)."],
    sheet_name: Annotated[str, "اسم الورقة."] = "Data",
    has_header: Annotated[str, "'auto' (افتراضي) / 'yes' / 'no' — هل أول صف عناوين؟"] = "auto",
    title: Annotated[str, "عنوان يُكتب فوق الجدول (اختياري)."] = "",
) -> str:
    """Turn raw pasted data (or a source file) into a clean, formatted Excel:
    auto-detected columns, bold frozen header, borders, auto-fit widths, numbers
    recognised (Arabic numerals too), and a right-to-left sheet when Arabic. This
    replaces a night of manual data entry with one call."""
    try:
        out = resolve_output_path(dest)
        src = data.strip()
        # Is `data` actually a path to a source file?
        if len(src) < 500 and os.path.exists(os.path.expanduser(src)) \
                and Path(src).suffix.lower() in (".csv", ".tsv", ".txt", ".pdf", ".docx"):
            ext = Path(src).suffix.lower()
            if ext in (".csv", ".tsv"):
                rows, header = _oc.csv_to_rows(os.path.expanduser(src)), True
            elif ext == ".txt":
                p = _oc.parse_tabular(Path(os.path.expanduser(src)).read_text(
                    encoding="utf-8", errors="replace"), has_header=has_header)
                rows, header = p["rows"], p["header"]
            elif ext == ".docx":
                tabs = _oc.docx_tables_to_rows(os.path.expanduser(src))
                if not tabs:
                    return "❌ لا يوجد جدول في ملف Word."
                rows, header = tabs[0], True
            else:  # pdf
                tabs = _oc.pdf_tables_to_rows(os.path.expanduser(src))
                if not tabs:
                    return "❌ تعذّر العثور على جدول في PDF."
                rows, header = tabs[0], True
        else:
            parsed = _oc.parse_tabular(data, has_header=has_header)
            rows, header = parsed["rows"], parsed["header"]

        if not rows:
            return "❌ لم أجد بيانات صالحة لتحويلها."
        path = _oc.rows_to_xlsx(rows, out, sheet_name=sheet_name, header=header,
                                title=title)
        n_rows = len(rows) - (1 if header else 0)
        n_cols = max(len(r) for r in rows)
        rtl = "نعم (RTL عربي)" if _oc.any_arabic(rows) else "لا"
        return (f"✅ تم إنشاء الإكسل: {path}\n"
                f"📊 {n_rows} صف × {n_cols} عمود · ترويسة: {'نعم' if header else 'لا'} · "
                f"عربي: {rtl}")
    except Exception as exc:
        return f"❌ data_to_excel: {exc}"


@tool
def office_convert(
    src: Annotated[str, "مسار الملف المصدر."],
    target_format: Annotated[str, "الصيغة الهدف بدون نقطة: xlsx / csv / pdf / docx."],
    dest: Annotated[str, "مسار الناتج (اختياري — نفس الاسم بامتداد جديد)."] = "",
) -> str:
    """Convert precisely between Office formats (Arabic & English):
    xlsx→csv/pdf/docx · csv/txt→xlsx · docx→xlsx/pdf · pdf→xlsx · pptx→pdf.
    Uses pure-Python engines (no LibreOffice needed for these), with correct
    Arabic shaping in generated PDFs and RTL sheets/tables."""
    try:
        src = os.path.expanduser(src)
        out = resolve_output_path(dest) if dest else None
        path = _oc.convert(src, target_format, out)
        size = os.path.getsize(path)
        return (f"✅ تم التحويل: {Path(src).name} → {path}\n"
                f"📦 الحجم: {round(size/1024, 1)} KB")
    except Exception as exc:
        return f"❌ office_convert: {exc}"


@tool
def extract_tables(
    src: Annotated[str, "ملف PDF أو Word يحوي جداول."],
    dest: Annotated[str, "مسار ملف الإكسل الناتج (.xlsx) — كل جدول في ورقة."],
) -> str:
    """Extract EVERY table from a PDF or Word file into one Excel workbook — each
    table on its own sheet, formatted and Arabic-aware. Ideal for pulling report
    tables into a spreadsheet without retyping."""
    try:
        from openpyxl import Workbook
        src = os.path.expanduser(src)
        ext = Path(src).suffix.lower()
        if ext == ".pdf":
            tables = _oc.pdf_tables_to_rows(src)
        elif ext == ".docx":
            tables = _oc.docx_tables_to_rows(src)
        else:
            return f"❌ الصيغة {ext} غير مدعومة — استخدم PDF أو Word."
        if not tables:
            return "❌ لم أجد أي جدول في الملف."

        out = resolve_output_path(dest)
        Path(out).parent.mkdir(parents=True, exist_ok=True)
        # Build a multi-sheet workbook by writing the first table then appending.
        from openpyxl import load_workbook
        first = tables[0]
        _oc.rows_to_xlsx(first, out, sheet_name="جدول 1", header=True)
        if len(tables) > 1:
            wb = load_workbook(out)
            for i, tbl in enumerate(tables[1:], start=2):
                ws = wb.create_sheet(title=f"جدول {i}")
                rtl = _oc.any_arabic(tbl)
                ws.sheet_view.rightToLeft = rtl
                for row in tbl:
                    ws.append([_oc.coerce_value(c) for c in row])
            wb.save(out)
        return f"✅ استُخرج {len(tables)} جدول إلى: {out}"
    except Exception as exc:
        return f"❌ extract_tables: {exc}"
