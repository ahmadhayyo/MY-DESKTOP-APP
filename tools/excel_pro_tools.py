"""
excel_pro_tools.py — Professional Excel operations (openpyxl).

Goes far beyond create/read/edit: real-world office formatting, formulas,
charts, conditional highlighting, column autofit, sheet management, and a
one-shot "styled report" builder. No Microsoft Office required.

Tools:
  • excel_format_range(path, cell_range, ...)   — bold/italic, colors, borders,
                                                  alignment, number format, font size
  • excel_set_formula(path, cell, formula)      — write a formula (=SUM(A1:A9) …)
  • excel_add_total_row(path, columns, label)   — append a SUM/AVG total row
  • excel_add_chart(path, chart_type, ...)      — bar/line/pie chart from a range
  • excel_autofit(path, sheet)                  — autofit all column widths
  • excel_freeze_header(path, sheet)            — freeze the top row
  • excel_highlight(path, cell_range, rule, ...)— conditional highlight by rule
  • excel_add_sheet(path, sheet_name, data)     — add a new worksheet (optional data)
  • excel_style_report(path, ...)               — one call: header style + autofit +
                                                  freeze + banded rows
"""
from __future__ import annotations

import json
from pathlib import Path

from langchain_core.tools import tool


# ── helpers ───────────────────────────────────────────────────────────────────
def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return None
    except ImportError:
        return "❌ openpyxl غير مثبّت. ثبّته بـ: pip install openpyxl"


def _open(path: str):
    from openpyxl import load_workbook
    p = Path(path)
    if not p.is_file():
        return None, None, f"❌ الملف غير موجود: {path}\n   أنشئه أولاً بـ excel_create()."
    try:
        wb = load_workbook(str(p))
        return wb, p, None
    except Exception as e:
        return None, None, f"❌ تعذّر فتح الملف: {e}"


def _ws(wb, sheet: str):
    if sheet and sheet in wb.sheetnames:
        return wb[sheet]
    return wb.active


def _argb(hex_str: str) -> str:
    """openpyxl wants 8-char ARGB (FF + RRGGBB)."""
    h = (hex_str or "").lstrip("#").strip().upper()
    if len(h) == 6:
        return "FF" + h
    if len(h) == 8:
        return h
    return "FFFFFF00"  # yellow fallback


# ═══════════════════════════════════════════════════════════════════════════════
@tool
def excel_format_range(
    path: str,
    cell_range: str,
    bold: bool = False,
    italic: bool = False,
    font_color: str = "",
    fill_color: str = "",
    font_size: int = 0,
    align: str = "",
    border: bool = False,
    number_format: str = "",
    sheet: str = "",
) -> str:
    """تنسيق نطاق خلايا في Excel (غامق، ألوان، حدود، محاذاة، تنسيق أرقام).

    Args:
        path: مسار الملف
        cell_range: النطاق مثل 'A1:D1' أو خلية واحدة 'B2'
        bold/italic: غامق/مائل
        font_color: لون الخط hex مثل 'FFFFFF' (أبيض)
        fill_color: لون الخلفية hex مثل '1F4E79' (أزرق)
        font_size: حجم الخط (0 = بدون تغيير)
        align: 'center' / 'right' / 'left'
        border: إضافة حدود رفيعة
        number_format: مثل '#,##0.00' للأرقام أو '0%' للنسب أو '#,##0 "ريال"'
        sheet: اسم الورقة (فارغ = النشطة)
    """
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        ws = _ws(wb, sheet)
        thin = Side(style="thin", color="FF999999")
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_map = {"center": "center", "right": "right", "left": "left"}
        count = 0
        for row in ws[cell_range]:
            cells = row if isinstance(row, tuple) else (row,)
            for cell in cells:
                f = cell.font
                cell.font = Font(
                    bold=bold or f.bold, italic=italic or f.italic,
                    color=_argb(font_color) if font_color else (f.color.rgb if f.color else None),
                    size=font_size or f.size,
                    name=f.name,
                )
                if fill_color:
                    cell.fill = PatternFill("solid", fgColor=_argb(fill_color))
                if align:
                    cell.alignment = Alignment(horizontal=align_map.get(align, "general"),
                                               vertical="center")
                if border:
                    cell.border = bd
                if number_format:
                    cell.number_format = number_format
                count += 1
        wb.save(str(p))
        return f"✅ نُسّق {count} خلية في النطاق {cell_range}."
    except Exception as e:
        return f"❌ خطأ في التنسيق: {e}"


@tool
def excel_set_formula(path: str, cell: str, formula: str, sheet: str = "") -> str:
    """كتابة صيغة في خلية. مثال: excel_set_formula(path, 'D10', '=SUM(D2:D9)').

    صيغ شائعة: =SUM(range) =AVERAGE(range) =COUNT(range) =MAX(range) =MIN(range)
               =IF(cond,a,b) =VLOOKUP(...) =A2*B2 =A2*0.15
    """
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        ws = _ws(wb, sheet)
        ws[cell] = formula if formula.startswith("=") else f"={formula}"
        wb.save(str(p))
        return f"✅ كُتبت الصيغة في {cell}: {ws[cell].value}"
    except Exception as e:
        return f"❌ خطأ في الصيغة: {e}"


@tool
def excel_add_total_row(
    path: str,
    columns: str,
    label: str = "الإجمالي",
    func: str = "SUM",
    sheet: str = "",
) -> str:
    """إضافة صف إجمالي أسفل البيانات بصيغ تلقائية لكل عمود محدد.

    Args:
        path: مسار الملف
        columns: أحرف الأعمدة المراد جمعها، مفصولة بفواصل، مثل 'B,C,D'
        label: نص التسمية في العمود الأول (افتراضي 'الإجمالي')
        func: 'SUM' أو 'AVERAGE' أو 'MAX' أو 'MIN' أو 'COUNT'
        sheet: اسم الورقة
    """
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        from openpyxl.styles import Font
        ws = _ws(wb, sheet)
        last = ws.max_row
        total_row = last + 1
        cols = [c.strip().upper() for c in columns.split(",") if c.strip()]
        # label in first column
        first_col = ws.cell(row=1, column=1).column_letter
        ws[f"{first_col}{total_row}"] = label
        ws[f"{first_col}{total_row}"].font = Font(bold=True)
        for col in cols:
            ws[f"{col}{total_row}"] = f"={func}({col}2:{col}{last})"
            ws[f"{col}{total_row}"].font = Font(bold=True)
        wb.save(str(p))
        return f"✅ أُضيف صف {func} في الصف {total_row} للأعمدة: {', '.join(cols)}."
    except Exception as e:
        return f"❌ خطأ: {e}"


@tool
def excel_add_chart(
    path: str,
    chart_type: str,
    data_range: str,
    categories_range: str = "",
    title: str = "",
    anchor: str = "H2",
    sheet: str = "",
) -> str:
    """إضافة رسم بياني داخل ورقة Excel من نطاق بيانات.

    Args:
        path: مسار الملف
        chart_type: 'bar' (أعمدة) / 'line' (خطي) / 'pie' (دائري)
        data_range: نطاق القيم مع صف العنوان، مثل 'B1:B10' أو 'B1:D10'
        categories_range: نطاق الفئات (المحور)، مثل 'A2:A10'
        title: عنوان الرسم
        anchor: الخلية التي يُلصق عندها الرسم (مثل 'H2')
        sheet: اسم الورقة
    """
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        ws = _ws(wb, sheet)
        ct = chart_type.lower()
        chart = {"bar": BarChart, "column": BarChart,
                 "line": LineChart, "pie": PieChart}.get(ct, BarChart)()
        if title:
            chart.title = title

        # Parse range "B1:D10" → min/max col/row
        from openpyxl.utils.cell import range_boundaries
        min_c, min_r, max_c, max_r = range_boundaries(data_range)
        data = Reference(ws, min_col=min_c, max_col=max_c, min_row=min_r, max_row=max_r)
        chart.add_data(data, titles_from_data=True)

        if categories_range:
            c1, r1, c2, r2 = range_boundaries(categories_range)
            cats = Reference(ws, min_col=c1, max_col=c2, min_row=r1, max_row=r2)
            chart.set_categories(cats)

        ws.add_chart(chart, anchor)
        wb.save(str(p))
        return f"✅ أُضيف رسم {chart_type} «{title}» عند {anchor}."
    except Exception as e:
        return f"❌ خطأ في الرسم البياني: {e}"


@tool
def excel_autofit(path: str, sheet: str = "") -> str:
    """ضبط عرض كل الأعمدة تلقائياً حسب أطول محتوى (يحسّن القراءة كثيراً)."""
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        ws = _ws(wb, sheet)
        widths: dict = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col = cell.column_letter
                    widths[col] = max(widths.get(col, 0), len(str(cell.value)))
        for col, w in widths.items():
            ws.column_dimensions[col].width = min(max(w + 2, 8), 60)
        wb.save(str(p))
        return f"✅ تم ضبط عرض {len(widths)} عمود تلقائياً."
    except Exception as e:
        return f"❌ خطأ: {e}"


@tool
def excel_freeze_header(path: str, sheet: str = "", rows: int = 1) -> str:
    """تجميد الصفوف العلوية (العناوين) لتبقى ظاهرة عند التمرير. rows=عدد الصفوف."""
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        ws = _ws(wb, sheet)
        ws.freeze_panes = f"A{rows + 1}"
        wb.save(str(p))
        return f"✅ تم تجميد أول {rows} صف."
    except Exception as e:
        return f"❌ خطأ: {e}"


@tool
def excel_highlight(
    path: str,
    cell_range: str,
    rule: str,
    value: str = "",
    color: str = "FFC7CE",
    sheet: str = "",
) -> str:
    """تمييز الخلايا التي تحقق شرطاً بلون (تنسيق شرطي).

    Args:
        cell_range: النطاق مثل 'C2:C100'
        rule: 'greater' / 'less' / 'equal' / 'between' / 'contains'
        value: القيمة للمقارنة (لـ between استخدم 'min,max')
        color: لون التمييز hex (افتراضي أحمر فاتح)
        sheet: اسم الورقة
    """
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.styles import PatternFill
        ws = _ws(wb, sheet)
        fill = PatternFill("solid", fgColor=_argb(color))
        op_map = {"greater": "greaterThan", "less": "lessThan",
                  "equal": "equal", "between": "between"}
        if rule == "contains":
            col = cell_range.split(":")[0]
            ws.conditional_formatting.add(
                cell_range,
                FormulaRule(formula=[f'ISNUMBER(SEARCH("{value}",{col}))'], fill=fill),
            )
        elif rule == "between":
            lo, hi = (value.split(",") + ["0", "0"])[:2]
            ws.conditional_formatting.add(
                cell_range, CellIsRule(operator="between", formula=[lo, hi], fill=fill))
        else:
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator=op_map.get(rule, "greaterThan"),
                           formula=[value or "0"], fill=fill))
        wb.save(str(p))
        return f"✅ طُبّق تمييز شرطي ({rule} {value}) على {cell_range}."
    except Exception as e:
        return f"❌ خطأ في التنسيق الشرطي: {e}"


@tool
def excel_add_sheet(path: str, sheet_name: str, data: str = "") -> str:
    """إضافة ورقة عمل جديدة (مع بيانات JSON اختيارية)."""
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        if sheet_name in wb.sheetnames:
            return f"ℹ️ الورقة «{sheet_name}» موجودة بالفعل."
        ws = wb.create_sheet(title=sheet_name)
        n = 0
        if data:
            rows = json.loads(data)
            if rows and isinstance(rows[0], dict):
                headers = list(rows[0].keys())
                ws.append(headers)
                for r in rows:
                    ws.append([r.get(h, "") for h in headers])
                n = len(rows)
            elif rows:
                for r in rows:
                    ws.append(list(r))
                n = len(rows)
        wb.save(str(p))
        return f"✅ أُضيفت ورقة «{sheet_name}»" + (f" ({n} صف)." if n else ".")
    except json.JSONDecodeError:
        return "❌ بيانات JSON غير صحيحة."
    except Exception as e:
        return f"❌ خطأ: {e}"


@tool
def excel_style_report(
    path: str,
    header_fill: str = "1F4E79",
    header_font_color: str = "FFFFFF",
    banded: bool = True,
    sheet: str = "",
) -> str:
    """تنسيق احترافي بنقرة واحدة: تنسيق صف العناوين + تجميده + ضبط الأعمدة + صفوف مخططة.

    مثالي لتحويل جدول بيانات خام إلى تقرير جاهز للعرض.

    Args:
        header_fill: لون خلفية العناوين hex
        header_font_color: لون خط العناوين hex
        banded: تلوين الصفوف بالتناوب لتسهيل القراءة
        sheet: اسم الورقة
    """
    err = _require_openpyxl()
    if err:
        return err
    wb, p, err = _open(path)
    if err:
        return err
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        ws = _ws(wb, sheet)
        max_col = ws.max_column
        max_row = ws.max_row
        thin = Side(style="thin", color="FFD0D0D0")
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        head_fill = PatternFill("solid", fgColor=_argb(header_fill))
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color=_argb(header_font_color), size=12)
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bd

        # Body: borders + banded rows
        band = PatternFill("solid", fgColor="FFF2F6FB")
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = bd
                if banded and r % 2 == 0:
                    cell.fill = band

        # Freeze header + autofit
        ws.freeze_panes = "A2"
        widths: dict = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    widths[cell.column_letter] = max(
                        widths.get(cell.column_letter, 0), len(str(cell.value)))
        for col, w in widths.items():
            ws.column_dimensions[col].width = min(max(w + 2, 10), 55)

        wb.save(str(p))
        return (f"✅ تقرير منسّق احترافياً: عناوين ملوّنة + مجمّدة + {max_col} عمود مضبوط"
                + (" + صفوف مخططة" if banded else "") + ".")
    except Exception as e:
        return f"❌ خطأ في التنسيق: {e}"
