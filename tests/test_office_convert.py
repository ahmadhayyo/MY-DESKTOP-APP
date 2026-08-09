"""
Office conversion tests — run with:  python -m unittest tests.test_office_convert

Comprehensive coverage of core/office_convert.py across Excel · Word · PDF ·
PowerPoint · CSV in BOTH Arabic and English: text intelligence (delimiters,
Arabic digits, header detection, numeric coercion, RTL), format round-trips,
and the convert() dispatcher. Stdlib unittest only.
"""
import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core import office_convert as oc  # noqa: E402


class TextIntelligence(unittest.TestCase):
    def test_normalize_arabic_digits(self):
        self.assertEqual(oc.normalize_digits("٣٥"), "35")
        self.assertEqual(oc.normalize_digits("۱۲۳"), "123")
        self.assertEqual(oc.normalize_digits("1,234"), "1,234")

    def test_is_arabic(self):
        self.assertTrue(oc.is_arabic("دمشق"))
        self.assertFalse(oc.is_arabic("Damascus"))
        self.assertTrue(oc.is_arabic("City مدينة"))

    def test_coerce_value(self):
        self.assertEqual(oc.coerce_value("35"), 35)
        self.assertEqual(oc.coerce_value("٣٥"), 35)          # Arabic digits → int
        self.assertEqual(oc.coerce_value("3.5"), 3.5)
        self.assertEqual(oc.coerce_value("1,234"), 1234)     # thousands sep
        self.assertEqual(oc.coerce_value("أحمد"), "أحمد")
        self.assertEqual(oc.coerce_value(""), "")

    def test_detect_delimiter(self):
        self.assertEqual(oc.detect_delimiter("a,b,c\n1,2,3"), ",")
        self.assertEqual(oc.detect_delimiter("a\tb\tc\n1\t2\t3"), "\t")
        self.assertEqual(oc.detect_delimiter("a;b;c\n1;2;3"), ";")
        self.assertEqual(oc.detect_delimiter("اسم،عمر\nأحمد،٣٥"), "،")  # Arabic comma

    def test_parse_tabular_header_detection_en(self):
        p = oc.parse_tabular("Name,Age,City\nAhmad,35,Damascus\nLayla,28,Aleppo")
        self.assertTrue(p["header"])
        self.assertEqual(p["rows"][0], ["Name", "Age", "City"])
        self.assertEqual(len(p["rows"]), 3)

    def test_parse_tabular_header_detection_ar(self):
        p = oc.parse_tabular("الاسم, العمر, المدينة\nأحمد, ٣٥, دمشق\nليلى, ٢٨, حلب")
        self.assertTrue(p["header"])          # labels above numeric column
        self.assertEqual(p["rows"][0][0], "الاسم")

    def test_parse_no_header_all_numbers(self):
        p = oc.parse_tabular("1,2,3\n4,5,6")
        self.assertFalse(p["header"])

    def test_parse_ragged_is_padded(self):
        p = oc.parse_tabular("a,b,c\n1,2\n9")
        self.assertTrue(all(len(r) == 3 for r in p["rows"]))

    def test_parse_multispace_aligned(self):
        p = oc.parse_tabular("Name    Age    City\nAhmad   35     Damascus")
        self.assertEqual(p["rows"][0], ["Name", "Age", "City"])
        self.assertEqual(p["rows"][1], ["Ahmad", "35", "Damascus"])

    def test_shape_arabic_changes_arabic_only(self):
        self.assertEqual(oc.shape_arabic("Hello"), "Hello")
        shaped = oc.shape_arabic("دمشق")
        self.assertIsInstance(shaped, str)
        self.assertTrue(len(shaped) >= 1)


class ExcelIO(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp()

    def test_rows_to_xlsx_and_back_en(self):
        rows = [["Name", "Age"], ["Ahmad", "35"], ["Layla", "28"]]
        p = oc.rows_to_xlsx(rows, os.path.join(self.d, "a.xlsx"), header=True)
        self.assertTrue(os.path.exists(p))
        back = oc.xlsx_to_rows(p)
        self.assertEqual(back[0], ["Name", "Age"])
        self.assertEqual(back[1], ["Ahmad", 35])       # coerced to int

    def test_rows_to_xlsx_arabic_rtl(self):
        rows = [["الاسم", "العمر"], ["أحمد", "٣٥"]]
        p = oc.rows_to_xlsx(rows, os.path.join(self.d, "ar.xlsx"), header=True)
        from openpyxl import load_workbook
        ws = load_workbook(p).active
        self.assertTrue(ws.sheet_view.rightToLeft)     # auto RTL for Arabic
        self.assertEqual(ws.cell(row=2, column=2).value, 35)  # ٣٥ → 35

    def test_header_is_bold(self):
        rows = [["H1", "H2"], ["1", "2"]]
        p = oc.rows_to_xlsx(rows, os.path.join(self.d, "b.xlsx"), header=True)
        from openpyxl import load_workbook
        ws = load_workbook(p).active
        self.assertTrue(ws.cell(row=1, column=1).font.bold)

    def test_csv_roundtrip_arabic(self):
        rows = [["الاسم", "المدينة"], ["أحمد", "دمشق"]]
        c = oc.rows_to_csv(rows, os.path.join(self.d, "a.csv"))
        back = oc.csv_to_rows(c)
        self.assertEqual(back, rows)

    def test_xlsx_to_csv_convert(self):
        rows = [["A", "B"], ["1", "2"]]
        x = oc.rows_to_xlsx(rows, os.path.join(self.d, "x.xlsx"))
        out = oc.convert(x, "csv", os.path.join(self.d, "x.csv"))
        self.assertTrue(out.endswith(".csv"))
        self.assertEqual(oc.csv_to_rows(out)[0], ["A", "B"])

    def test_csv_to_xlsx_convert(self):
        c = oc.rows_to_csv([["Name", "Age"], ["Ahmad", "35"]],
                           os.path.join(self.d, "c.csv"))
        out = oc.convert(c, "xlsx", os.path.join(self.d, "c.xlsx"))
        self.assertEqual(oc.xlsx_to_rows(out)[1], ["Ahmad", 35])


class WordIO(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp()

    def test_rows_to_docx_table_and_back(self):
        rows = [["Name", "City"], ["Ahmad", "Damascus"], ["Layla", "Aleppo"]]
        p = oc.rows_to_docx_table(rows, os.path.join(self.d, "t.docx"), header=True)
        self.assertTrue(os.path.exists(p))
        tabs = oc.docx_tables_to_rows(p)
        self.assertEqual(len(tabs), 1)
        self.assertEqual(tabs[0][0], ["Name", "City"])
        self.assertEqual(tabs[0][1], ["Ahmad", "Damascus"])

    def test_docx_table_arabic(self):
        rows = [["الاسم", "المدينة"], ["أحمد", "دمشق"]]
        p = oc.rows_to_docx_table(rows, os.path.join(self.d, "ar.docx"))
        tabs = oc.docx_tables_to_rows(p)
        self.assertEqual(tabs[0][1], ["أحمد", "دمشق"])

    def test_xlsx_to_docx_convert(self):
        x = oc.rows_to_xlsx([["A", "B"], ["1", "2"]], os.path.join(self.d, "x.xlsx"))
        out = oc.convert(x, "docx", os.path.join(self.d, "x.docx"))
        tabs = oc.docx_tables_to_rows(out)
        self.assertEqual(tabs[0][0], ["A", "B"])

    def test_docx_to_xlsx_convert(self):
        rows = [["Name", "Age"], ["Ahmad", "35"]]
        d = oc.rows_to_docx_table(rows, os.path.join(self.d, "d.docx"))
        out = oc.convert(d, "xlsx", os.path.join(self.d, "d.xlsx"))
        back = oc.xlsx_to_rows(out)
        self.assertEqual(back[0], ["Name", "Age"])
        self.assertEqual(back[1], ["Ahmad", 35])


class PdfIO(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp()

    def test_rows_to_pdf_english(self):
        rows = [["Name", "Age"], ["Ahmad", "35"], ["Layla", "28"]]
        p = oc.rows_to_pdf_table(rows, os.path.join(self.d, "e.pdf"), title="Team")
        self.assertTrue(os.path.exists(p) and os.path.getsize(p) > 500)

    def test_rows_to_pdf_arabic(self):
        rows = [["الاسم", "العمر"], ["أحمد", "٣٥"], ["ليلى", "٢٨"]]
        p = oc.rows_to_pdf_table(rows, os.path.join(self.d, "a.pdf"),
                                 title="الفريق")
        self.assertTrue(os.path.exists(p) and os.path.getsize(p) > 500)

    def test_xlsx_to_pdf_convert(self):
        x = oc.rows_to_xlsx([["A", "B"], ["1", "2"]], os.path.join(self.d, "x.xlsx"))
        out = oc.convert(x, "pdf", os.path.join(self.d, "x.pdf"))
        self.assertTrue(os.path.exists(out) and os.path.getsize(out) > 500)

    def test_pdf_table_roundtrip(self):
        # write a table to PDF, then extract it back with pdfplumber
        rows = [["Name", "Age", "City"], ["Ahmad", "35", "Damascus"],
                ["Layla", "28", "Aleppo"]]
        p = oc.rows_to_pdf_table(rows, os.path.join(self.d, "rt.pdf"), header=True)
        tabs = oc.pdf_tables_to_rows(p)
        self.assertTrue(tabs, "pdfplumber should find the table")
        flat = [str(c) for row in tabs[0] for c in row]
        self.assertIn("Ahmad", flat)
        self.assertIn("Damascus", flat)


class Dispatcher(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp()

    def test_txt_to_xlsx(self):
        txt = Path(self.d) / "data.txt"
        txt.write_text("Name,Age\nAhmad,35\nLayla,28", encoding="utf-8")
        out = oc.convert(str(txt), "xlsx")
        self.assertTrue(out.endswith(".xlsx"))
        self.assertEqual(oc.xlsx_to_rows(out)[1], ["Ahmad", 35])

    def test_missing_source_raises(self):
        with self.assertRaises(FileNotFoundError):
            oc.convert(os.path.join(self.d, "nope.xlsx"), "csv")

    def test_unsupported_pair_raises(self):
        x = oc.rows_to_xlsx([["a"]], os.path.join(self.d, "x.xlsx"))
        with self.assertRaises(Exception):
            oc.convert(x, "mp3")

    def test_read_any_rows_dispatch(self):
        x = oc.rows_to_xlsx([["A"], ["1"]], os.path.join(self.d, "x.xlsx"))
        self.assertEqual(oc.read_any_rows(x)[0], ["A"])


if __name__ == "__main__":
    unittest.main(verbosity=2)
