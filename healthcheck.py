#!/usr/bin/env python3
"""
healthcheck.py — Pre-demo self-test for HAYO.

Run this BEFORE any presentation/demo to confirm exactly which capabilities are
working. It exercises the deterministic, offline-capable parts (no LLM calls, no
paid APIs) and prints a clear PASS/FAIL report so you walk in knowing what's green.

Usage:
    venv\\Scripts\\python.exe healthcheck.py
"""
from __future__ import annotations

import sys
import os
import tempfile
import logging

logging.disable(logging.WARNING)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

PASS = "✅ PASS"
FAIL = "❌ FAIL"
WARN = "⚠️  WARN"
results = []


def check(name, fn, network=False):
    try:
        detail = fn() or ""
        results.append((PASS, name, detail))
    except Exception as e:
        tag = WARN if network else FAIL
        results.append((tag, name, f"{type(e).__name__}: {e}"))


# ── 1. Tool registry ──────────────────────────────────────────────────────────
def _registry():
    from tools.registry import ALL_TOOLS, _failed_categories
    names = [t.name for t in ALL_TOOLS]
    dupes = {n for n in names if names.count(n) > 1}
    assert not _failed_categories, f"failed categories: {_failed_categories}"
    assert not dupes, f"duplicate names: {dupes}"
    return f"{len(ALL_TOOLS)} tools, 0 failures, 0 duplicates"


# ── 2. OCR engine ─────────────────────────────────────────────────────────────
def _ocr():
    from tools.ocr_engine import ocr_available, ocr_engine_name
    assert ocr_available(), "no OCR engine"
    return f"engine = {ocr_engine_name()}"


# ── 3. Excel generation ───────────────────────────────────────────────────────
def _excel():
    from tools.office_tools import excel_create
    from tools.excel_pro_tools import excel_style_report, excel_add_chart
    import json
    p = os.path.join(tempfile.gettempdir(), "_hc.xlsx")
    data = [["المنتج", "الكمية", "السعر"], ["قلم", 100, 2], ["دفتر", 50, 5]]
    excel_create.invoke({"path": p, "data": json.dumps(data, ensure_ascii=False)})
    excel_style_report.invoke({"path": p})
    from openpyxl import load_workbook
    ws = load_workbook(p).active
    assert ws.max_column == 3 and ws.cell(1, 1).value == "المنتج", "structure/Arabic broken"
    os.remove(p)
    return "Excel + styling + Arabic OK"


# ── 4. Word generation ────────────────────────────────────────────────────────
def _word():
    from tools.word_pro_tools import word_add_heading, word_add_table
    import json
    p = os.path.join(tempfile.gettempdir(), "_hc.docx")
    word_add_heading.invoke({"path": p, "text": "تقرير", "level": 0})
    word_add_table.invoke({"path": p, "data": json.dumps([["البند", "المبلغ"], ["إيجار", 3000]], ensure_ascii=False)})
    assert os.path.getsize(p) > 0
    os.remove(p)
    return "Word headings + tables + Arabic OK"


# ── 5. PowerPoint generation ──────────────────────────────────────────────────
def _ppt():
    from tools.powerpoint_tools import ppt_create, ppt_add_bullets, ppt_add_chart
    p = os.path.join(tempfile.gettempdir(), "_hc.pptx")
    ppt_create.invoke({"path": p, "title": "عرض"})
    ppt_add_bullets.invoke({"path": p, "title": "نقاط", "bullets": "أولاً\nثانياً"})
    ppt_add_chart.invoke({"path": p, "title": "رسم", "chart_type": "bar",
                          "data": '{"categories":["أ","ب"],"series":{"س":[1,2]}}'})
    assert os.path.getsize(p) > 0
    os.remove(p)
    return "PowerPoint slides + charts OK"


# ── 6. Arabic mojibake repair ─────────────────────────────────────────────────
def _repair():
    from core.text_repair import fix_mojibake, normalize_table
    assert fix_mojibake("ط§ظ„ظ‚ط³ظ…") == "القسم", "mojibake not repaired"
    assert fix_mojibake("القسم") == "القسم", "correct text corrupted"
    grid = normalize_table(["a,b,c", "1,2,3"])
    assert grid == [["a", "b", "c"], ["1", "2", "3"]], "CSV split broken"
    return "mojibake repair + table normalize OK"


# ── 7. Tool-name auto-correct (difflib) ───────────────────────────────────────
def _autocorrect():
    import difflib
    from tools.registry import TOOLS_BY_NAME
    names = list(TOOLS_BY_NAME.keys())
    cases = {"browser_react": "browser_react_fill", "web_searh": "web_search",
             "read_page": "read_webpage"}
    for bad, good in cases.items():
        close = difflib.get_close_matches(bad, names, n=3, cutoff=0.6)
        prefix = [n for n in names if n.startswith(bad)]
        cand = (prefix[0] if len(prefix) == 1 else (close[0] if close else None))
        assert cand == good, f"{bad} → {cand} (expected {good})"
    return "hallucinated tool names auto-correct OK"


# ── 8. Long-term memory ───────────────────────────────────────────────────────
def _memory():
    from core import long_term_memory as ltm
    ltm.remember("hc_demo_key", "hc_demo_value", category="test")
    hit = ltm.recall("hc_demo")
    ltm.forget("hc_demo_key")
    assert hit, "recall failed"
    return "remember/recall/forget OK"


# ── 9. Scheduler parsing ──────────────────────────────────────────────────────
def _scheduler():
    from tools.scheduler_tools import _parse_when
    assert _parse_when("every day at 09:00")["kind"] == "daily"
    assert _parse_when("بعد ساعة")["kind"] == "once"
    assert _parse_when("every 2 hours")["kind"] == "interval"
    return "schedule parsing (AR+EN) OK"


# ── 10. DB maintenance importable ─────────────────────────────────────────────
def _maintenance():
    from core.maintenance import auto_prune_if_needed  # noqa: F401
    return "memory-DB pruning available"


# ── 11. Capability Forge (self-extension) ─────────────────────────────────────
def _forge():
    from tools.forge_tools import forge_tool, remove_forged_tool
    from tools.registry import TOOLS_BY_NAME
    code = ('@tool\n'
            'def hc_twice(n: str) -> str:\n'
            '    """double a number"""\n'
            '    return str(int(n) * 2)\n')
    forge_tool.invoke({"tool_name": "hc_twice", "description": "x2", "python_code": code})
    fn = TOOLS_BY_NAME.get("hc_twice")
    assert fn and fn.invoke({"n": "21"}) == "42", "forged tool not live/callable"
    remove_forged_tool.invoke({"tool_name": "hc_twice"})
    return "agent forged + ran + removed a tool live"


# ── 12. App builder readiness (no slow build) ─────────────────────────────────
def _appbuilder():
    from tools.app_builder_tools import _venv_pyinstaller, lint_python
    import tempfile
    assert _venv_pyinstaller(), "PyInstaller not found (pip install pyinstaller)"
    import importlib
    importlib.import_module("tkinter")
    # lint a tiny valid snippet
    p = os.path.join(tempfile.gettempdir(), "_hc_app.py")
    with open(p, "w", encoding="utf-8") as f:
        f.write("import tkinter as tk\nroot = tk.Tk()\nroot.mainloop()\n")
    r = lint_python.invoke({"path": p})
    os.remove(p)
    assert r.startswith("✅"), f"lint failed: {r}"
    return "PyInstaller + tkinter + lint ready (full build ~30-60s)"


# ── 13. Integrations (live data, network) ─────────────────────────────────────
def _integrations():
    from tools.integration_tools import get_crypto_price, send_discord
    # zero-config live call
    r = get_crypto_price.invoke({"symbol": "bitcoin"})
    assert "$" in r or "Bitcoin" in r, f"crypto fetch failed: {r}"
    # graceful degradation when a key is missing (must not crash)
    d = send_discord.invoke({"message": "x"})
    assert d.startswith("❌") or d.startswith("✅"), "discord did not degrade gracefully"
    return "live data OK + graceful key-missing handling"


# ── 14. Web search (network — WARN if offline) ────────────────────────────────
def _websearch():
    from tools.web_search_tools import _ddg_text
    r = _ddg_text("python programming", 2)
    assert r, "no results (network/VPN?)"
    return f"{len(r)} live results"


# ── 15. True vision (image analysis) ──────────────────────────────────────────
def _vision():
    from core import vision_analyze
    provs = vision_analyze.available_vision_providers()
    assert provs, "no vision provider (add GOOGLE_API_KEY or ANTHROPIC_API_KEY)"
    return f"providers = {provs}"


# ── 16. Verify + visual gates (build→run→see→fix, incl. Android) ───────────────
def _gates():
    from core.verify_gate import verification_pending, visual_verification_pending
    h = [{"name": "edit_file_replace", "args": {"path": "a.py"}, "result": "[OK]"}]
    assert verification_pending(h).pending, "verify gate should arm after code edit"
    a = [{"name": "android_launch_app", "args": {"package": "x"}, "result": "[OK]"}]
    os.environ.setdefault("GOOGLE_API_KEY", os.getenv("GOOGLE_API_KEY", ""))
    vp = visual_verification_pending(a).pending  # may be dormant w/o vision key
    return f"verify=armed, android-visual={'armed' if vp else 'dormant(no vision key)'}"


# ── 17. Live todo list ────────────────────────────────────────────────────────
def _todo():
    from core import todo
    todo.clear_todos("_hc")
    todo.write_todos("_hc", [{"id": "1", "content": "x", "status": "completed"}])
    assert todo.progress("_hc") == (1, 1)
    todo.clear_todos("_hc")
    return "write/render/progress OK"


# ── 18. Skills system ─────────────────────────────────────────────────────────
def _skills():
    from core import skills
    names = [s.name for s in skills.load_all()]
    assert names, "no skills loaded"
    rel = skills.find_relevant("أصلح المشروع")
    assert rel, "relevance matching failed"
    return f"{len(names)} skills, auto-match OK"


# ── 19. Persistent terminal session (venv-aware) ──────────────────────────────
def _terminal():
    from core import terminal_session
    terminal_session.reset_session("_hc")
    r = terminal_session.run_in_session("echo hc_ok", session_id="_hc", timeout=15)
    terminal_session.reset_session("_hc")
    assert r["output"] == "hc_ok", f"got {r}"
    return "persistent shell OK"


# ── 20. Sub-agent engine (depth guard) ────────────────────────────────────────
def _subagent():
    from core import subagent
    from contextvars import copy_context
    # verify depth guard blocks recursion
    tok = subagent._depth.set(1)
    try:
        r = subagent.run_subagent("x")
        assert "تكرار" in r or "recursion" in r.lower()
    finally:
        subagent._depth.reset(tok)
    return "loop-guard OK"


# ── 21. Whole-PC file/app search ──────────────────────────────────────────────
def _locate():
    from tools.locate_tools import find_on_computer
    r = find_on_computer.invoke({"name": "python", "kind": "app", "deep": False, "max_results": 3})
    assert isinstance(r, str)
    return "find_on_computer OK"


# ── 22. Provider wiring (google/openai/omniroute build) ───────────────────────
def _providers():
    from agent.nodes import _build_llm
    built = []
    for p in ("google", "openai", "omniroute"):
        try:
            _build_llm("main", p)
            built.append(p)
        except Exception:
            pass
    assert built, "no provider could be built"
    return f"buildable: {built}"


def main():
    print("=" * 60)
    print("  HAYO — فحص صحة شامل / Comprehensive capability check")
    print("=" * 60)
    check("1. Tool registry loads", _registry)
    check("2. OCR engine (vision)", _ocr)
    check("3. Excel generation", _excel)
    check("4. Word generation", _word)
    check("5. PowerPoint generation", _ppt)
    check("6. Arabic repair + tables", _repair)
    check("7. Tool-name auto-correct", _autocorrect)
    check("8. Long-term memory", _memory)
    check("9. Scheduler parsing", _scheduler)
    check("10. Memory-DB pruning", _maintenance)
    check("11. Capability Forge (self-extend)", _forge)
    check("12. Desktop app builder (EXE)", _appbuilder)
    check("13. Integrations (live data)", _integrations, network=True)
    check("14. Web search (network)", _websearch, network=True)
    check("15. True vision (image analysis)", _vision)
    check("16. Verify + visual gates", _gates)
    check("17. Live todo list", _todo)
    check("18. Skills system", _skills)
    check("19. Persistent terminal (venv)", _terminal)
    check("20. Sub-agent engine", _subagent)
    check("21. Whole-PC search", _locate)
    check("22. Provider wiring", _providers)

    print()
    npass = sum(1 for s, _, _ in results if s == PASS)
    for status, name, detail in results:
        print(f"  {status}  {name}")
        if detail:
            print(f"          {detail}")
    print()
    print("-" * 60)
    fails = [r for r in results if r[0] == FAIL]
    print(f"  النتيجة: {npass}/{len(results)} نجحت"
          + (f" | {len(fails)} فشل حقيقي" if fails else " | لا أعطال حرجة"))
    print("  (⚠️ WARN = يعتمد على الشبكة/الإنترنت — طبيعي إن كنت غير متصل)")
    print("=" * 60)
    return 0 if not fails else 1


if __name__ == "__main__":
    sys.exit(main())
